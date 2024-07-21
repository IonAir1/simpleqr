from bs4 import BeautifulSoup
from PIL import Image
from openpyxl.utils import get_column_letter
import pypdfium2 as pdfium
import openpyxl
import openpyxl.drawing
import openpyxl.drawing.image
import os.path
import shutil


def clean_string(string):
    return string.string.replace('\n','').replace('Ã‘','ñ').replace('Ã±','ñ').strip()


def to_pdf(file):
    if not os.path.exists("temp"):
        os.makedirs("temp")
    
    pdf = pdfium.PdfDocument(file)
    image = pdf[0].render(scale=4).to_pil()


    return image


def toDictionary(html, file_path):
    info = html.find_all("td")
    student = {}
    
    student["number"] = clean_string(info[0])
    student["name"] = clean_string(info[1])
    student["sex"] = clean_string(info[2])
    student["course"] = clean_string(info[4])
    student["email"] = clean_string(info[5])
    student["mobile"] = clean_string(info[6])
    
    img = html.find("img")

    html_path = file_path[:file_path.rfind('/')]
    path = html_path + img['src'].replace('./','/')
    student["picture"] = path

    return student


#get students' info
def scrape_html(html):
    with open(html, "r", encoding='utf8') as f:
        doc = BeautifulSoup(f, "html.parser")
    students_html = doc.find_all("tr")
    students_html.pop(0) #remove header

    students = []
    total_students = len(students_html)
    for student in students_html:
        print("Scraping HTML File ({}/{})".format(students_html.index(student)+1,total_students))
        students.append(toDictionary(student, html))
    return students


def create_workbook(students, **kwargs):
    number = kwargs.get('number', False) 
    name = kwargs.get('name', True)
    sex = kwargs.get('sex', False)
    course = kwargs.get('course', False)
    email = kwargs.get('email', False)
    mobile = kwargs.get('mobile', False)
    picture = kwargs.get('picture', True)

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    current_column = 1
    space_before_pic = kwargs.get('space_before_pic', 0)

    first= True
    total_students = len(students)
    for student in students:
        i = students.index(student)

        print("Constructing Excel File ({}/{})".format(i+1, total_students))

        if number:
            ws.cell(row=i+1, column=current_column).value = student['number']
            if first:
                ws.column_dimensions[get_column_letter(current_column)].width = 15
            current_column += 1

        if name:
            ws.cell(row=i+1, column=current_column).value = student['name']
            if first:
                ws.column_dimensions[get_column_letter(current_column)].width = 30
            current_column += 1
            
        if sex:
            ws.cell(row=i+1, column=current_column).value = student['sex']
            if first:
                ws.column_dimensions[get_column_letter(current_column)].width = 5
            current_column += 1

        if course:
            ws.cell(row=i+1, column=current_column).value = student['course']
            if first:
                ws.column_dimensions[get_column_letter(current_column)].width = 25
            current_column += 1

        if email:
            ws.cell(row=i+1, column=current_column).value = student['email']
            if first:
                ws.column_dimensions[get_column_letter(current_column)].width = 25
            current_column += 1

        if mobile:
            ws.cell(row=i+1, column=current_column).value = student['mobile']
            if first:
                ws.column_dimensions[get_column_letter(current_column)].width = 18
            current_column += 1

        if picture:
            current_column += space_before_pic

            if first:
                ws.column_dimensions[get_column_letter(current_column)].width = 50
            cell = get_column_letter(current_column) + str(i+1)
            ws.row_dimensions[i+1].height = 200

            if student['picture'].endswith('.pdf'):
                if kwargs.get('skip_pdf', False) == False:
                    to_pdf(student['picture']).save("temp/temp{}.png".format(i))
                    img = openpyxl.drawing.image.Image("temp/temp{}.png".format(i))
                    img.width = 200
                    img.height = 200
                    ws.add_image(img, cell)
            elif not student['picture'].endswith("saved_resource"):
                img = openpyxl.drawing.image.Image(student['picture'])
                img.width = 200
                img.height = 200
                ws.add_image(img, cell)


        current_column = 1
        first = False

    return wb


def html_to_excel(file_path, output_path, **kwargs):
    print("Opening HTML File")
    students = scrape_html(file_path)
    wb = create_workbook(students, **kwargs)
    # wb = create_workbook(students, number=True, sex=True, course=True, email=True, mobile=True)
    print("Saving Excel File")
    wb.save(output_path)
    if os.path.exists("temp"):
        shutil.rmtree("temp")
        os.makedirs("temp")
    if kwargs.get('open', True) == True:
        os.startfile(output_path)